"""
pdf_to_excel.py
Extract product name and CAS numbers from SDS (Safety Data Sheet) PDFs.

Output columns: Chemical Name (= product trade name from SDS Section 1)
                CAS Number  (all CAS numbers from Section 3, duplicates kept)

Usage:
  Single PDF : python pdf_to_excel.py input.pdf output.xlsx
  Folder     : python pdf_to_excel.py folder/ [output.xlsx]
"""

import re
import os
import sys
import math
from typing import List, Dict

import pandas as pd
import pdfplumber
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Regex patterns ────────────────────────────────────────────────────────────
CAS_RE = re.compile(r'\b(\d{1,7}-\d{2}-\d)\b')
CAS_LABEL_RE = re.compile(
    r'\bCAS\s*(?:No\.?|Number|#)?\s*[:\s]\s*(\d{1,7}-\d{2}-\d)',
    re.IGNORECASE
)
SECTION3_RE = re.compile(
    r'SECTION\s*3\b|3\s*[.:]\s*COMPOSITION|COMPOSITION.*INGREDIENT',
    re.IGNORECASE
)
SECTION4_RE = re.compile(
    r'SECTION\s*4\b|4\s*[.:]\s*FIRST[\s\-]?AID',
    re.IGNORECASE
)
NOISE_RE = re.compile(
    r'^(EC\b|EINECS|REACH|Registration|Concentration|Hazard|Classification|'
    r'Section|Page|Safety|Revision|Version|Index|Regulation|Percentage|%|'
    r'Component\b|Compound\b|Ingredient\b|Mixture\b|Composition\b|'
    r'Annex|\d+[\s/])',
    re.IGNORECASE
)
HCODE_RE = re.compile(r'\bH\d{3}\b')
_NC = r'[A-Za-z0-9 \t,\.\(\)\-/\[\]]'

# ── Product name extraction from Section 1 ───────────────────────────────────

# Ordered list of (pattern, group_index) to try on each page text
_PRODUCT_NAME_PATTERNS = [
    # "1.1 Product identifier: Name"   (ASUBLANC, T079360116, CORAFIX)
    # Use [ \t]* — NOT \s* — to avoid consuming the newline when the name is absent
    re.compile(r'1\.1\.?\s*Product\s*identifier\s*[:\-][ \t]*([^\n]{3,120})', re.IGNORECASE),
    # "Trade name : Name" or "Trade name - Name"   (Caustic Soda, GGFix, BEMACRON)
    re.compile(r'Trade\s*[Nn]ame\s*[:\-]\s*([^\n]{3,120})', re.IGNORECASE),
    # "Trade name\n<Name on next line>"   (Drimaren Orange)
    re.compile(r'Trade\s*[Nn]ame\s*\n\s*([^\n]{3,120})', re.IGNORECASE),
    # "Product Name Corazol ..." (CORAZOL — no colon; name immediately follows)
    re.compile(r'Product\s*[Nn]ame\s+([A-Za-z][^\n]{2,100})', re.IGNORECASE),
    # "PRODUCT NAME: Name" or "PRODUCT NAME Name"   (Binder SC-20C)
    re.compile(r'PRODUCT\s+NAME\s*:?\s*([^\n]{3,80})', re.IGNORECASE),
    # "Chemical Name : Name"   (Hydrogen Peroxide)
    re.compile(r'Chemical\s+[Nn]ame\s*[:\-]\s*([^\n]{3,80})', re.IGNORECASE),
    # "Substance name : Name ..."   (fallback for Hydrogen Peroxide)
    re.compile(r'Substance\s+[Nn]ame\s*[:\-]\s*([^\n]{3,80})', re.IGNORECASE),
]
# Noise sub-strings to cut off captured product names
_NAME_STOP = re.compile(
    r'\s{3,}|Synonyms\b|Other means\b|CAS\s*No|EC\s*No|Material number|'
    r'SDS NUMBER|CAT NO',
    re.IGNORECASE
)


def _clean_product_name(raw: str) -> str:
    """Trim trailing noise from a product name candidate."""
    name = _NAME_STOP.split(raw)[0].strip()
    name = re.sub(r'\s+', ' ', name).strip(' |,-')
    return name


def extract_product_name(pdf) -> str:
    """
    Extract the trade/commercial name from Section 1 of an SDS PDF.

    Handles:
    - "1.1 Product identifier: Name"        (ASUBLANC, T079360116)
    - "Trade name : Name"                   (Caustic Soda, GGFix, BEMACRON)
    - "Trade name\\nName"                   (Drimaren Orange)
    - "Product Name Corazol ..."            (CORAZOL)
    - "PRODUCT NAME: Name"                  (Binder SC-20C)
    - "Chemical Name : Name"               (Hydrogen Peroxide)
    - Name appears BEFORE a blank "Product Name:" label  (Machine-Oil)
    """
    for page in pdf.pages[:3]:
        text = page.extract_text() or ''
        lines = text.splitlines()

        # ── Special case: name sits on the line BEFORE an empty "Product Name:" ──
        # Used by Machine-Oil-MSDS.pdf where:
        #   Lily White/Crystal Clear Sewing Machine Oil
        #   Product Name:
        for i, line in enumerate(lines):
            if re.match(r'\s*Product\s*Name\s*:\s*$', line, re.IGNORECASE):
                for j in range(i - 1, max(i - 5, -1), -1):
                    candidate = lines[j].strip()
                    if candidate and len(candidate) > 3 and not re.search(
                        r'safety|sheet|section|company|identification|msds|sds',
                        candidate, re.IGNORECASE
                    ):
                        return _clean_product_name(candidate)
                break

        # ── Standard pattern matching ──────────────────────────────────────────
        for pat in _PRODUCT_NAME_PATTERNS:
            m = pat.search(text)
            if m:
                name = _clean_product_name(m.group(1))
                if (len(name) > 2 and not re.search(
                    r'^\d{4}|^version|^revision|^page\b|^section\b|^date\b|'
                    r'^non-applicable|^n/?a\b',
                    name, re.IGNORECASE
                )):
                    return name

    return ''


# ── Section 3 page detection ──────────────────────────────────────────────────

def find_section3_pages(pdf) -> List[int]:
    """Return 0-based page indices belonging to the Section 3 block."""
    pages: List[int] = []
    in_s3 = False
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ''
        if SECTION3_RE.search(text):
            in_s3 = True
        if in_s3:
            pages.append(i)
            if SECTION4_RE.search(text):
                break
    return pages


# ── Table-based CAS extraction ────────────────────────────────────────────────

def _trim_name(raw: str) -> str:
    name = raw.split('\n')[0]
    name = re.sub(r'\s+(Self-classified|ATP\s+\w+).*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*;?\s*H\d{3}.*$', '', name)
    return name.strip()


def find_name_for_cas(text: str, cas: str) -> str:
    """
    Look for an ingredient/component name near `cas` in raw text.

    Checks (in order):
    1. Text on the SAME line before the CAS number
    2. Lines ABOVE the CAS line (up to 5 lines back), skipping label-only lines
       e.g. 'White mineral oil (petroleum)' sits 2 lines above 'CAS number : 8042-47-5'
    3. Wrapped name: if the line above ends with AND/&, append the line below the CAS
       e.g. 'PARAFFIN WAXES AND' / '8002-74-2 ...' / 'HYDROCARBON WAXES'
    """
    _SKIP = re.compile(
        r'^(CAS|EC\b|REACH|Index|Reg|Revision|Version|Page|Section|'
        r'Concentration|Prepared|Note|\d)',
        re.IGNORECASE
    )

    lines = text.splitlines()
    for j, line in enumerate(lines):
        if cas not in line:
            continue

        # 1. Same-line prefix
        idx = line.index(cas)
        before = re.sub(r'[\d\s\.,;%\-/()]+$', '', line[:idx]).strip()
        if len(before) >= 3 and not _SKIP.search(before):
            return _trim_name(before)

        # 2 & 3. Look backwards up to 5 lines
        for k in range(j - 1, max(j - 6, -1), -1):
            candidate = lines[k].strip()
            if not candidate:
                continue
            if _SKIP.search(candidate):
                continue
            if CAS_RE.search(candidate):
                break  # hit another CAS row — stop looking up

            name = re.sub(r'[\d\s\.,;%\-/()]+$', '', candidate).strip()
            if len(name) < 3:
                continue

            # 3. Wrapped: name fragment ends with AND / & — glue the line after CAS
            if re.search(r'\b(AND|&)\s*$', name, re.IGNORECASE):
                after = lines[j + 1].strip() if j + 1 < len(lines) else ''
                after = re.sub(r'[\d\s\.,;%\-/()]+$', '', after).strip()
                if after and not CAS_RE.search(after) and not _SKIP.search(after):
                    return _trim_name(name + ' ' + after)

            return _trim_name(name)

    return ''


def _find_cas_header(table):
    for hi, row in enumerate(table[:8]):
        row_text = ' '.join(clean(c) for c in row if c)
        if not re.search(r'\bCAS\b', row_text, re.IGNORECASE):
            continue
        non_empty = [clean(c) for c in row if c]
        if non_empty and all(re.match(r'CAS\s*:', v, re.IGNORECASE) for v in non_empty):
            continue
        name_col = cas_col = -1
        for ci, cell in enumerate(row):
            ct = clean(cell).lower()
            if re.search(r'\bcas\b', ct):
                cas_col = ci
            if re.search(r'\b(name|substance|chemical|component|ingredi)\b', ct):
                name_col = ci
        if cas_col >= 0:
            return hi, name_col, cas_col
    return -1, -1, -1


def clean(x) -> str:
    if x is None:
        return ''
    return re.sub(r'\s+', ' ', str(x)).strip()


def extract_cas_from_table(table) -> List[tuple]:
    """Extract (cas, chemical_name) pairs from a PDF table. Duplicates preserved."""
    if not table:
        return []

    results: List[tuple] = []

    # ── Format C: 'Identification' column with embedded CAS label ─────────────
    if table[0]:
        id_col = name_col = -1
        for ci, cell in enumerate(table[0]):
            ct = clean(cell)
            if re.search(r'\bidentif', ct, re.IGNORECASE):
                id_col = ci
            if re.search(r'\b(chemical\s*name|substance\s*name|component)\b', ct, re.IGNORECASE):
                name_col = ci
        if id_col >= 0:
            for row in table[1:]:
                if not any(c for c in row if c):
                    continue
                if id_col >= len(row) or not row[id_col]:
                    continue
                cell_text = clean(row[id_col])
                cas_nums = CAS_LABEL_RE.findall(cell_text) or CAS_RE.findall(cell_text)
                chem_name = ''
                if name_col >= 0 and name_col < len(row) and row[name_col]:
                    chem_name = _trim_name(clean(row[name_col]))
                for cas in cas_nums:
                    results.append((cas, chem_name))
            return results

    # ── Format A/B: explicit CAS column in header ─────────────────────────────
    hi, name_col, cas_col = _find_cas_header(table)
    if hi < 0:
        return []

    for row in table[hi + 1:]:
        if not any(c for c in row if c):
            continue
        cas_nums = []
        if 0 <= cas_col < len(row) and row[cas_col]:
            cas_nums = CAS_RE.findall(clean(row[cas_col]))
        else:
            for cell in row:
                cas_nums.extend(CAS_RE.findall(clean(cell) if cell else ''))
        chem_name = ''
        if name_col >= 0 and name_col < len(row) and row[name_col]:
            chem_name = _trim_name(clean(row[name_col]))
        for cas in cas_nums:
            results.append((cas, chem_name))

    return results


# ── Text-based CAS extraction ─────────────────────────────────────────────────

def extract_cas_from_text(text: str, exclude: set = None) -> List[tuple]:
    """
    Extract (cas, chemical_name) pairs from plain page text.
    CAS numbers already in `exclude` are skipped.

    P1 – Same-line CAS label:  'Sodium Hydroxide CAS No: 1310-73-2 >99%'
    P2 – CAS label on its own line
    P3 – Plain-text table with 'COMPONENT  CAS NO.' header
    P4 – Bare 'Name  CAS_NUMBER'
    """
    if exclude is None:
        exclude = set()

    found: List[tuple] = []
    seen: set = set(exclude)

    def add(cas: str, name: str = ''):
        if cas not in seen:
            seen.add(cas)
            found.append((cas, name.strip()))

    lines = text.splitlines()

    # P1: 'Name CAS No: XXXX' on same line
    for m in re.finditer(
        rf'({_NC}{{2,80}}?)'
        r'\s+CAS\s*(?:No\.?|Number|#)?\s*[:\s]\s*(\d{1,7}-\d{2}-\d)',
        text, re.IGNORECASE
    ):
        add(m.group(2), m.group(1))

    # P2: CAS label on its own line (no adjacent name)
    for line in lines:
        m = CAS_LABEL_RE.search(line)
        if m:
            add(m.group(1), '')

    # P3: plain-text table — 'COMPONENT CAS NO.' header
    header_idx = comp_pos = cas_pos = -1
    for i, line in enumerate(lines):
        if re.search(
                r'\bCOMPONENT\b.*\bCAS\b|\bCHEMICAL[\s\-]NAME\b.*\bCAS\b',
                line, re.IGNORECASE):
            header_idx = i
            mc = re.search(r'\bCOMPONENT\b|\bCHEMICAL\b', line, re.IGNORECASE)
            mk = re.search(r'\bCAS\b', line, re.IGNORECASE)
            if mc and mk:
                comp_pos, cas_pos = mc.start(), mk.start()
            break

    if header_idx >= 0 and 0 <= comp_pos < cas_pos:
        for line in lines[header_idx + 1:]:
            if not line.strip():
                continue
            if re.search(r'^(SECTION\s*\d|PREPARED|Note)', line, re.IGNORECASE):
                break
            for m in CAS_RE.finditer(line):
                cas = m.group(1)
                name_part = line[:m.start()].strip()
                add(cas, name_part)

    # P4: bare 'Name  CAS_NUMBER'
    for m in re.finditer(
        rf'([A-Z]{_NC}{{9,80}}?)\s+(\d{{1,7}}-\d{{2}}-\d)\b',
        text
    ):
        name = m.group(1).strip()
        if not re.search(r'section|page|version|revision', name, re.IGNORECASE):
            if not HCODE_RE.search(name):
                add(m.group(2), name)

    return found


# ── Per-PDF entry point ───────────────────────────────────────────────────────

def extract_from_pdf(pdf_path: str) -> Dict:
    """
    Extract product name and all CAS numbers from a single SDS PDF.
    Returns {'product_name': str, 'cas_list': List[str]}

    Table extraction strategy:
    - Per page: use the FIRST table that yields at least one CAS number.
      This avoids pdfplumber detecting the same composition table multiple times
      (overlapping extractions of main table, M-factor table, concentration table).
    - Cross-page dedup: if the same CAS was already extracted from an earlier
      page, skip it.  This handles multi-page table continuations (ASUBLANC).
    - Within-page preservation: CAS that appears multiple times inside one
      table (e.g. SC-20C POLYACRYLATE + POLYACRYLATE THICKENER, both 25035-69-2)
      is kept, because `seen_cross_page` is only updated AFTER the current page
      loop finishes.

    Text fallback: only applied when zero CAS were found from any table.
    """
    items: List[Dict] = []                # each: {'cas': str, 'chem_name': str}
    seen_cross_page: set = set()          # tracks unique CAS seen in previous pages

    with pdfplumber.open(pdf_path) as pdf:
        product_name = extract_product_name(pdf)

        s3_pages = find_section3_pages(pdf)
        scan_indices = s3_pages if s3_pages else list(range(len(pdf.pages)))

        # ── Table extraction ──────────────────────────────────────────────────
        for pi in scan_indices:
            page_pairs: List[tuple] = []
            for t in (pdf.pages[pi].extract_tables() or []):
                t_pairs = extract_cas_from_table(t)
                if t_pairs:
                    page_pairs = t_pairs
                    break   # first qualifying table per page wins

            # Preserve within-page duplicates; skip cross-page duplicates.
            for cas, chem_name in page_pairs:
                if cas not in seen_cross_page:
                    items.append({'cas': cas, 'chem_name': chem_name})
            seen_cross_page.update(cas for cas, _ in page_pairs)

        # ── Text fallback (only when tables yielded nothing) ──────────────────
        if not items:
            for pi in scan_indices:
                text = pdf.pages[pi].extract_text() or ''
                for cas, chem_name in extract_cas_from_text(text):
                    if cas not in seen_cross_page:
                        seen_cross_page.add(cas)
                        items.append({'cas': cas, 'chem_name': chem_name})

        # ── Fill in missing chemical names from page text ──────────────────────
        if any(not item['chem_name'] for item in items):
            page_texts = [
                pdf.pages[pi].extract_text() or ''
                for pi in scan_indices
            ]
            full_text = '\n'.join(page_texts)
            for item in items:
                if not item['chem_name']:
                    item['chem_name'] = find_name_for_cas(full_text, item['cas'])

    return {'product_name': product_name, 'items': items}


def autofit_worksheet(worksheet, min_width: int = 10, max_width: int = 60) -> None:
    column_widths: Dict[int, int] = {}
    header_fill = PatternFill(fill_type='solid', start_color='D9E1F2', end_color='D9E1F2')
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    for col_idx in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = Font(bold=True)
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = header_border

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            value = '' if cell.value is None else str(cell.value)
            longest_line = max((len(part) for part in value.splitlines()), default=0)
            current = column_widths.get(cell.column, 0)
            if longest_line > current:
                column_widths[cell.column] = longest_line

    effective_widths: Dict[int, int] = {}
    for col_idx, longest in column_widths.items():
        width = max(min_width, min(longest + 2, max_width))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        effective_widths[col_idx] = width

    for row_idx in range(1, worksheet.max_row + 1):
        max_lines = 1
        for col_idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            text = '' if value is None else str(value)
            width = max(1, int(effective_widths.get(col_idx, min_width)) - 2)
            wrapped_lines = 0
            for line in text.splitlines() or ['']:
                wrapped_lines += max(1, math.ceil(len(line) / width))
            if wrapped_lines > max_lines:
                max_lines = wrapped_lines
        worksheet.row_dimensions[row_idx].height = min(120, max(15, max_lines * 15))


# ── Folder processing & Excel output ─────────────────────────────────────────

def process_folder(folder: str, output_xlsx: str) -> pd.DataFrame:
    pdf_files = sorted(
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith('.pdf')
    )

    rows: List[Dict] = []
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        print(f'Processing: {filename}')
        try:
            result = extract_from_pdf(pdf_path)
            product_name = result['product_name'] or filename
            items = result['items']

            if items:
                for item in items:
                    rows.append({'Product Name': product_name,
                                 'Chemical Name': item['chem_name'],
                                 'CAS Number': item['cas']})
            else:
                rows.append({'Product Name': product_name,
                             'Chemical Name': '',
                             'CAS Number': 'N/A'})
        except Exception as exc:
            print(f'  ERROR: {exc}')
            rows.append({'Product Name': filename, 'Chemical Name': '', 'CAS Number': f'ERROR: {exc}'})

    df = pd.DataFrame(rows, columns=['Product Name', 'Chemical Name', 'CAS Number'])
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Chemicals')
        autofit_worksheet(writer.sheets['Chemicals'])

    print(f'\nSaved {len(df)} rows  →  {output_xlsx}')
    return df


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) == 3 and sys.argv[1].lower().endswith('.pdf'):
        pdf_path, output = sys.argv[1], sys.argv[2]
        result = extract_from_pdf(pdf_path)
        product_name = result['product_name'] or os.path.basename(pdf_path)
        rows = [{'Product Name': product_name,
                 'Chemical Name': item['chem_name'],
                 'CAS Number': item['cas']}
                for item in result['items']] or \
               [{'Product Name': product_name, 'Chemical Name': '', 'CAS Number': 'N/A'}]
        df = pd.DataFrame(rows, columns=['Product Name', 'Chemical Name', 'CAS Number'])
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Chemicals')
            autofit_worksheet(writer.sheets['Chemicals'])
        print(f'Saved {len(df)} rows  →  {output}')
        print(df.to_string(index=False))

    elif len(sys.argv) >= 2:
        folder = sys.argv[1]
        output = (sys.argv[2] if len(sys.argv) > 2
                  else os.path.join(folder, 'chemicals_output.xlsx'))
        df = process_folder(folder, output)
        print()
        print(df.to_string(index=False))

    else:
        print('Usage:')
        print('  Single PDF : python pdf_to_excel.py input.pdf output.xlsx')
        print('  Folder     : python pdf_to_excel.py folder/ [output.xlsx]')
        sys.exit(1)


if __name__ == '__main__':
    main()
